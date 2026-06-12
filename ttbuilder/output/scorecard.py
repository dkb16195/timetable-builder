"""Timetable quality scorecard.

Implements the metric set proposed in research_raw/metrics.md (the XHSTT /
ITC2007-grounded 11-metric scorecard, §4) as far as the solved-timetable data
allows. Metric names cite the M-numbers used there. Reads only
configs/*.yaml + assignments.csv. Emits scorecard.md and scorecard.xlsx.
"""
from __future__ import annotations

import os
import statistics
from collections import defaultdict
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .excel import Context, load_context

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")


@dataclass
class Metric:
    name: str          # cites metrics.md M-number
    value: str
    direction: str     # "lower better" / "higher better" / "target band"
    interpretation: str


def _teacher_day_positions(ctx: Context):
    """For each (teacher, week, day): sorted indices of busy teaching periods.

    Busy = teaching assignment OR pinned commitment (XHSTT idle-time
    definition counts any busy time — metrics.md §2.1).
    """
    grid = ctx.grid
    teach_idx = {d: {p: i for i, p in enumerate(grid.teaching(d))}
                 for d in grid.days}
    out: dict[tuple, list[int]] = defaultdict(list)
    for tid, slots in ctx.teacher_slots.items():
        for (w, d, p) in slots:
            out[(tid, w, d)].append(teach_idx[d][p])
    for tid, pins in ctx.pinned.items():
        for (w, d, p) in pins:
            idx = teach_idx.get(d, {}).get(p)
            if idx is not None and idx not in out[(tid, w, d)]:
                out[(tid, w, d)].append(idx)
    return {k: sorted(v) for k, v in out.items()}


def compute_metrics(ctx: Context) -> tuple[list[Metric], dict]:
    grid = ctx.grid
    metrics: list[Metric] = []
    raw: dict = {}

    # ---- M1 Teacher Gap Rate + M10 Compactness + days of attendance ------
    positions = _teacher_day_positions(ctx)
    total_idle = total_busy = total_span = 0
    per_teacher_idle: dict[str, int] = defaultdict(int)
    days_present: dict[str, int] = defaultdict(int)
    for (tid, _w, _d), pos in positions.items():
        span = pos[-1] - pos[0] + 1
        idle = span - len(pos)
        total_idle += idle
        total_busy += len(pos)
        total_span += span
        per_teacher_idle[tid] += idle
        days_present[tid] += 1
    n_teachers = len({k[0] for k in positions})
    taught_slots = sum(len(s) for s in ctx.teacher_slots.values())
    gap_rate = total_idle / taught_slots if taught_slots else 0.0
    worst_tid, worst_gaps = ("", 0)
    if per_teacher_idle:
        worst_tid, worst_gaps = max(per_teacher_idle.items(), key=lambda kv: kv[1])
    tname = {t["id"]: t["name"] for t in ctx.teachers}
    mean_gaps_per_week = (total_idle / n_teachers / len(grid.weeks)
                          if n_teachers else 0.0)
    raw["gap_total"] = total_idle
    raw["gap_rate"] = gap_rate
    metrics.append(Metric(
        "M1 Teacher Gap Rate",
        f"{total_idle} idle periods over the fortnight; rate {gap_rate:.3f}; "
        f"mean {mean_gaps_per_week:.2f}/teacher/week; worst "
        f"{tname.get(worst_tid, worst_tid)} ({worst_gaps})",
        "lower better",
        "Free periods stranded between two busy periods in a day (XHSTT "
        "idle-time definition, pinned commitments count as busy)."))

    compact = total_busy / total_span if total_span else 1.0
    raw["span_efficiency"] = compact
    metrics.append(Metric(
        "M10 Teacher Compactness (Span Efficiency)",
        f"{compact:.3f} (busy {total_busy} / span {total_span})",
        "higher better",
        "Share of each teacher's daily first-to-last span actually taught; "
        "1.0 means no embedded gaps anywhere."))

    mean_days = (sum(days_present.values()) / n_teachers / len(grid.weeks)
                 if n_teachers else 0.0)
    metrics.append(Metric(
        "Schedule Compactness — days of attendance (metrics.md §2.2)",
        f"mean {mean_days:.2f} busy days/teacher/week",
        "lower better",
        "ClusterBusyTimes-style count of days each teacher must be present; "
        "fewer days = more compressible part-time schedules."))

    # ---- M3 Subject Spread + fortnight balance ---------------------------
    sec_week_day: dict[tuple, int] = defaultdict(int)
    sec_week: dict[tuple, int] = defaultdict(int)
    for sec, smap in ctx.by_section.items():
        for (w, d, _p), ms in smap.items():
            sec_week_day[(sec, w, d)] += len(ms)
            sec_week[(sec, w)] += len(ms)
    sections = set(ctx.by_section)
    bad_secs = {sec for (sec, w, d), n in sec_week_day.items() if n > 1}
    spread_ok = 100.0 * (len(sections) - len(bad_secs)) / len(sections)
    raw["spread_ok_pct"] = spread_ok
    metrics.append(Metric(
        "M3 Subject Spread Score — within-week",
        f"{spread_ok:.1f}% of sections never meet twice on one day",
        "higher better",
        "Distributed-practice spread (SpreadEventsConstraint, max 1/day): "
        "the share of sections whose meetings land on distinct days."))

    balanced = sum(
        1 for sec in sections
        if abs(sec_week.get((sec, grid.weeks[0]), 0)
               - sec_week.get((sec, grid.weeks[-1]), 0)) <= 1)
    balance_pct = 100.0 * balanced / len(sections)
    raw["week_balance_pct"] = balance_pct
    metrics.append(Metric(
        "M3 Subject Spread Score — fortnight A/B balance",
        f"{balance_pct:.1f}% of sections have |week A − week B| ≤ 1 meetings",
        "higher better",
        "Two-week-cycle balance term from metrics.md: a 4-in-A/1-in-B split "
        "defeats spaced practice even if each week looks fine."))

    # ---- M4 Same-day repeats ---------------------------------------------
    repeats = sum(1 for n in sec_week_day.values() if n > 1)
    raw["same_day_repeats"] = repeats
    metrics.append(Metric(
        "M4 Same-Day Repeats",
        f"{repeats} section-days with ≥2 meetings of the same section",
        "lower better",
        "Unplanned second lesson of a subject on one day; intentional "
        "doubles are not flagged in assignments.csv, so any configured "
        "doubles are included in this count."))

    # ---- M11 Room frequency ----------------------------------------------
    slots = grid.teaching_slots()
    rooms = {r["id"] for r in ctx.cfg.get("rooms", [])} | set(ctx.by_room)
    used = sum(len(ctx.by_room.get(r, {})) for r in rooms)
    avail = len(rooms) * len(slots)
    freq = 100.0 * used / avail if avail else 0.0
    raw["room_util_pct"] = freq
    in_band = "inside" if 60 <= freq <= 85 else "outside"
    metrics.append(Metric(
        "M11 Specialist Room Frequency (room utilisation %)",
        f"{freq:.1f}% of room×teaching-slot capacity used "
        f"({used}/{avail}; {in_band} the 60–85% target band)",
        "target band 60–85%",
        "SMG frequency rate; 100% is NOT the goal — schools need slack for "
        "cover, exams and clubs."))

    # ---- M7 Teacher room stability ----------------------------------------
    teacher_rooms: dict[str, set] = defaultdict(set)
    teacher_day_rooms: dict[tuple, set] = defaultdict(set)
    for tid, tmap in ctx.by_teacher.items():
        for (w, d, _p), ms in tmap.items():
            for m in ms:
                if m.room:
                    teacher_rooms[tid].add(m.room)
                    teacher_day_rooms[(tid, w, d)].add(m.room)
    itc_penalty = sum(len(r) - 1 for r in teacher_rooms.values())
    mean_rooms_day = (statistics.mean(len(r) for r in teacher_day_rooms.values())
                      if teacher_day_rooms else 0.0)
    raw["teacher_rooms_per_day"] = mean_rooms_day
    metrics.append(Metric(
        "M7 Teacher Room Stability",
        f"mean {mean_rooms_day:.2f} distinct rooms/teacher/day; "
        f"ITC2007 penalty Σ(distinct rooms − 1) = {itc_penalty}",
        "lower better",
        "Fewer rooms per teacher per day = less corridor time and a "
        "stable base room (ITC2007 RoomStability)."))

    # ---- Class room stability (M7 companion / ITC2007 on sections) --------
    sec_rooms: dict[str, set] = defaultdict(set)
    for sec, smap in ctx.by_section.items():
        for ms in smap.values():
            for m in ms:
                if m.room:
                    sec_rooms[sec].add(m.room)
    mean_sec_rooms = (statistics.mean(len(r) for r in sec_rooms.values())
                      if sec_rooms else 0.0)
    one_room_pct = (100.0 * sum(1 for r in sec_rooms.values() if len(r) == 1)
                    / len(sec_rooms)) if sec_rooms else 0.0
    raw["class_rooms_mean"] = mean_sec_rooms
    metrics.append(Metric(
        "Class Room Stability (ITC2007 RoomStability on sections)",
        f"mean {mean_sec_rooms:.2f} distinct rooms/section; "
        f"{one_room_pct:.1f}% of sections stay in a single room",
        "lower better",
        "A class that always meets in the same room loses less lesson time "
        "to setup and student movement (metrics.md §2.6/§2.9)."))

    # ---- M5 Teacher daily load balance ------------------------------------
    loads: dict[tuple, int] = defaultdict(int)
    for tid, slots_t in ctx.teacher_slots.items():
        for (w, d, _p) in slots_t:
            loads[(tid, w, d)] += 1
    per_teacher: dict[str, list[int]] = defaultdict(list)
    for (tid, _w, _d), n in loads.items():
        per_teacher[tid].append(n)
    stdevs = [statistics.pstdev(v) for v in per_teacher.values() if v]
    cvs = [statistics.pstdev(v) / statistics.mean(v)
           for v in per_teacher.values() if v and statistics.mean(v) > 0]
    mean_std = statistics.mean(stdevs) if stdevs else 0.0
    mean_cv = statistics.mean(cvs) if cvs else 0.0
    raw["load_balance_std"] = mean_std
    metrics.append(Metric(
        "M5 Teacher Daily Load Balance",
        f"mean per-teacher stdev of periods/working-day = {mean_std:.2f} "
        f"(mean CV {mean_cv:.2f})",
        "lower better",
        "Evenness of each teacher's load across the days they work "
        "(LimitBusyTimes-style); high values mean crammed days next to "
        "near-empty ones."))

    # ---- First/last-period distribution (diagnostic, metrics.md §4) -------
    first_n = last_n = total_n = 0
    per_teacher_edge: dict[str, int] = defaultdict(int)
    for tid, slots_t in ctx.teacher_slots.items():
        for (w, d, p) in slots_t:
            day_teach = grid.teaching(d)
            total_n += 1
            edge = False
            if p == day_teach[0]:
                first_n += 1
                edge = True
            if p == day_teach[-1]:
                last_n += 1
                edge = True
            if edge:
                per_teacher_edge[tid] += 1
    edge_counts = [per_teacher_edge.get(t["id"], 0) for t in ctx.teachers
                   if t["id"] in ctx.teacher_slots]
    fair_spread = (max(edge_counts) - min(edge_counts)) if edge_counts else 0
    first_pct = 100.0 * first_n / total_n if total_n else 0.0
    last_pct = 100.0 * last_n / total_n if total_n else 0.0
    raw["first_pct"], raw["last_pct"] = first_pct, last_pct
    metrics.append(Metric(
        "First/Last-Period Distribution (diagnostic)",
        f"{first_pct:.1f}% of teaching slots in the first period, "
        f"{last_pct:.1f}% in the last; teacher fairness spread "
        f"(max−min first+last count) = {fair_spread}",
        "balanced better",
        "Edge-of-day teaching should be shared fairly across staff; a wide "
        "spread means a few teachers absorb all the P1/last-period slots."))

    return metrics, raw


def build_scorecard(config_path: str, assignments_path: str,
                    out_dir: str) -> dict:
    os.makedirs(out_dir, exist_ok=True)
    ctx = load_context(config_path, assignments_path)
    metrics, raw = compute_metrics(ctx)

    md_path = os.path.join(out_dir, "scorecard.md")
    lines = [
        f"# Timetable Quality Scorecard — {ctx.title}",
        "",
        "Metric set per `research_raw/metrics.md` (XHSTT / ITC2007 grounded; "
        "M-numbers cite its §4 proposal). Computed from the solved timetable "
        "(`assignments.csv`) only; clash-freeness is a gate handled by the "
        "validator, not scored here.",
        "",
        "| Metric | Value | Direction | Interpretation |",
        "|---|---|---|---|",
    ]
    for m in metrics:
        lines.append(f"| {m.name} | {m.value} | {m.direction} | "
                     f"{m.interpretation} |")
    lines.append("")
    with open(md_path, "w") as f:
        f.write("\n".join(lines))

    wb = Workbook()
    ws = wb.active
    ws.title = "Scorecard"
    ws.cell(1, 1, f"{ctx.title} — Timetable Quality Scorecard").font = \
        Font(bold=True, size=12)
    heads = ["Metric", "Value", "Direction", "Interpretation"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(2, j, h)
        c.font = Font(bold=True, size=9)
        c.fill = HEADER_FILL
    for i, m in enumerate(metrics, 3):
        ws.cell(i, 1, m.name).font = Font(bold=True, size=8)
        ws.cell(i, 2, m.value).font = Font(size=8)
        ws.cell(i, 3, m.direction).font = Font(size=8)
        ws.cell(i, 4, m.interpretation).font = Font(size=8)
    for col, w in zip("ABCD", (44, 60, 18, 80)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    wb.save(os.path.join(out_dir, "scorecard.xlsx"))
    return raw
