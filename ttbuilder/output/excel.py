"""Excel output layer.

Reads ONLY the canonical artifacts (the school config YAML + assignments.csv
+ enrolments.csv) — never solver internals — and renders the seven workbooks
described in PLAN.md: master, department, year-group, teacher, per-grade
student timetables, room utilisation and staff loading.

Entry point: generate_all(config_path, assignments_path, enrolments_path,
out_dir).
"""
from __future__ import annotations

import os
import re
from collections import defaultdict
from dataclasses import dataclass, field

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styles --
GREY = PatternFill("solid", fgColor="DDDDDD")
DARKGREY = PatternFill("solid", fgColor="BFBFBF")
AMBER = PatternFill("solid", fgColor="FFE699")
RED = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")

TITLE_FONT = Font(bold=True, size=11)
HEAD_FONT = Font(bold=True, size=8)
CELL_FONT = Font(size=6)
PIN_FONT = Font(size=6, italic=True, color="808080")
LABEL_FONT = Font(bold=True, size=7)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="BBBBBB")
MEDIUM = Side(style="medium", color="555555")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DAY_END = Border(left=THIN, right=MEDIUM, top=THIN, bottom=THIN)

_SHEET_BAD = re.compile(r"[\[\]:*?/\\]")


def _sheet_name(name: str, used: set[str]) -> str:
    base = _SHEET_BAD.sub("-", str(name)).strip()[:31] or "Sheet"
    cand, n = base, 2
    while cand in used:
        suffix = f" ({n})"
        cand = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(cand)
    return cand


# ----------------------------------------------------------------- model --
@dataclass
class Grid:
    weeks: list[str]
    days: list[str]
    period_order: list[str]               # config order, all kinds
    kind: dict[str, str]                  # period id -> kind
    day_periods: dict[str, list[str]]     # day -> period ids (overrides applied)
    times: dict[tuple[str, str], tuple[str, str]]  # (day, pid) -> (start, end)

    def teaching(self, day: str) -> list[str]:
        return [p for p in self.day_periods[day] if self.kind[p] == "teaching"]

    def teaching_slots(self) -> list[tuple[str, str, str]]:
        return [(w, d, p) for w in self.weeks for d in self.days
                for p in self.teaching(d)]


@dataclass
class Meeting:
    section: str
    subject: str
    teacher_names: str
    room: str
    grade: str


@dataclass
class Context:
    cfg: dict
    grid: Grid
    title: str                             # "School — Year"
    teachers: list[dict]                   # config staffing teachers, name-sorted
    tier_max: dict[str, int]
    pinned: dict[str, dict[tuple[str, str, str], str]]
    df: pd.DataFrame
    by_teacher: dict[str, dict[tuple, list[Meeting]]] = field(default_factory=dict)
    by_section: dict[str, dict[tuple, list[Meeting]]] = field(default_factory=dict)
    by_grade: dict[str, dict[tuple, list[str]]] = field(default_factory=dict)
    by_room: dict[str, dict[tuple, list[str]]] = field(default_factory=dict)
    teacher_slots: dict[str, set[tuple]] = field(default_factory=dict)
    section_meta: dict[str, dict] = field(default_factory=dict)


def load_grid(cfg: dict) -> Grid:
    g = cfg["grid"]
    period_order = [p["id"] for p in g["periods"]]
    kind = {p["id"]: p["kind"] for p in g["periods"]}
    base_times = {p["id"]: (str(p["start"]), str(p["end"])) for p in g["periods"]}
    day_periods = {d: list(period_order) for d in g["days"]}
    for day, plist in (g.get("day_overrides") or {}).items():
        day_periods[day] = list(plist)
    times: dict[tuple[str, str], tuple[str, str]] = {}
    fri = g.get("friday_times") or {}
    for d in g["days"]:
        for p in day_periods[d]:
            t = base_times[p]
            if d == "Friday" and p in fri:
                t = (str(fri[p]["start"]), str(fri[p]["end"]))
            times[(d, p)] = t
    return Grid(weeks=list(g["weeks"]), days=list(g["days"]),
                period_order=period_order, kind=kind,
                day_periods=day_periods, times=times)


def _parse_slot_ref(ref: str) -> tuple[str, str, str]:
    w, d, p = ref.split(".")
    return w, d, p


def load_context(config_path: str, assignments_path: str) -> Context:
    with open(config_path) as f:
        cfg = yaml.safe_load(f)
    grid = load_grid(cfg)
    school = cfg.get("school", {})
    title = f"{school.get('name', 'School')} — {school.get('academic_year', '')}".strip(" —")
    staffing = cfg.get("staffing", {})
    teachers = sorted(staffing.get("teachers", []), key=lambda t: t["name"])
    tier_max = dict(staffing.get("tiers", {}))
    pinned: dict[str, dict[tuple, str]] = defaultdict(dict)
    for pb in cfg.get("pinned_busy", []) or []:
        label = pb.get("label", "pinned")
        for ref in pb.get("slots", []):
            pinned[pb["teacher"]][_parse_slot_ref(ref)] = label

    df = pd.read_csv(assignments_path, dtype=str).fillna("")
    ctx = Context(cfg=cfg, grid=grid, title=title, teachers=teachers,
                  tier_max=tier_max, pinned=dict(pinned), df=df)

    by_teacher: dict[str, dict] = defaultdict(lambda: defaultdict(list))
    by_section: dict[str, dict] = defaultdict(lambda: defaultdict(list))
    by_grade: dict[str, dict] = defaultdict(lambda: defaultdict(list))
    by_room: dict[str, dict] = defaultdict(lambda: defaultdict(list))
    teacher_slots: dict[str, set] = defaultdict(set)
    section_meta: dict[str, dict] = {}
    for r in df.itertuples(index=False):
        slot = (r.week, r.day, r.period)
        m = Meeting(section=r.section_id, subject=r.subject,
                    teacher_names=r.teacher_names, room=r.room, grade=r.grade)
        by_section[r.section_id][slot].append(m)
        meta = section_meta.setdefault(
            r.section_id, {"subject": r.subject, "grade": r.grade,
                           "teacher_names": set(), "meetings": 0})
        meta["meetings"] += 1
        if r.teacher_names:
            meta["teacher_names"].update(r.teacher_names.split("|"))
        if r.grade:
            by_grade[r.grade][slot].append(r.section_id)
        if r.room:
            by_room[r.room][slot].append(r.section_id)
        for tid in (r.teachers.split("|") if r.teachers else []):
            by_teacher[tid][slot].append(m)
            teacher_slots[tid].add(slot)
    ctx.by_teacher = {k: dict(v) for k, v in by_teacher.items()}
    ctx.by_section = {k: dict(v) for k, v in by_section.items()}
    ctx.by_grade = {k: dict(v) for k, v in by_grade.items()}
    ctx.by_room = {k: dict(v) for k, v in by_room.items()}
    ctx.teacher_slots = dict(teacher_slots)
    ctx.section_meta = section_meta
    return ctx


def _short(label: str, n: int = 60) -> str:
    return label if len(label) <= n else label[: n - 1] + "…"


def assigned_per_week(ctx: Context, tid: str) -> float:
    return len(ctx.teacher_slots.get(tid, ())) / max(len(ctx.grid.weeks), 1)


def pinned_per_week(ctx: Context, tid: str) -> float:
    return len(ctx.pinned.get(tid, ())) / max(len(ctx.grid.weeks), 1)


# ------------------------------------------------------- master timetable --
def build_master(ctx: Context, path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    grid = ctx.grid
    for week in grid.weeks:
        ws = wb.create_sheet(f"Week {week}")
        ws.cell(1, 1, f"{ctx.title} — Master Timetable — Week {week}").font = TITLE_FONT
        # header rows: 2 = day (merged), 3 = period
        col = 2
        ws.cell(3, 1, "Teacher").font = HEAD_FONT
        day_cols: list[tuple[str, str, int]] = []
        for day in grid.days:
            pids = grid.day_periods[day]
            start_col = col
            for i, pid in enumerate(pids):
                c = ws.cell(3, col, pid)
                c.font = HEAD_FONT
                c.alignment = CENTER
                c.fill = HEADER_FILL if grid.kind[pid] == "teaching" else GREY
                c.border = DAY_END if i == len(pids) - 1 else BOX
                w = 11 if grid.kind[pid] == "teaching" else 4
                ws.column_dimensions[get_column_letter(col)].width = w
                day_cols.append((day, pid, col))
                col += 1
            ws.merge_cells(start_row=2, start_column=start_col,
                           end_row=2, end_column=col - 1)
            dc = ws.cell(2, start_col, day)
            dc.font = HEAD_FONT
            dc.alignment = CENTER
            dc.fill = HEADER_FILL
        ws.column_dimensions["A"].width = 18
        row = 4
        for t in ctx.teachers:
            tid = t["id"]
            name_cell = ws.cell(row, 1, t["name"])
            name_cell.font = LABEL_FONT
            name_cell.border = BOX
            tmeet = ctx.by_teacher.get(tid, {})
            tpin = ctx.pinned.get(tid, {})
            for day, pid, c in day_cols:
                cell = ws.cell(row, c)
                cell.border = DAY_END if pid == grid.day_periods[day][-1] else BOX
                if grid.kind[pid] != "teaching":
                    cell.fill = GREY
                    continue
                slot = (week, day, pid)
                ms = tmeet.get(slot)
                if ms:
                    cell.value = "\n+ ".join(
                        f"{m.section}\n{m.subject}\n{m.room}" for m in ms)
                    cell.font = CELL_FONT
                    cell.alignment = WRAP
                elif slot in tpin:
                    cell.value = _short(tpin[slot], 40)
                    cell.font = PIN_FONT
                    cell.alignment = WRAP
                    cell.fill = GREY
            ws.row_dimensions[row].height = 26
            row += 1
        ws.freeze_panes = "B4"
    wb.save(path)


# -------------------------------------------------- department timetables --
def build_departments(ctx: Context, path: str) -> None:
    subj_dept = {s["id"]: s.get("department", "Other")
                 for s in ctx.cfg.get("subjects", [])}
    dept_sections: dict[str, list[str]] = defaultdict(list)
    for sec, meta in ctx.section_meta.items():
        dept_sections[subj_dept.get(meta["subject"], "Other")].append(sec)
    grid = ctx.grid
    slots = grid.teaching_slots()
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for dept in sorted(dept_sections):
        ws = wb.create_sheet(_sheet_name(dept, used))
        ws.cell(1, 1, f"{ctx.title} — Department Timetable — {dept}").font = TITLE_FONT
        heads = ["Section", "Subject", "Grade", "Teachers"]
        for i, h in enumerate(heads, 1):
            c = ws.cell(2, i, h)
            c.font = HEAD_FONT
            c.fill = HEADER_FILL
        col_of: dict[tuple, int] = {}
        for j, (w, d, p) in enumerate(slots, len(heads) + 1):
            c = ws.cell(2, j, f"{w} {d[:3]} {p}")
            c.font = HEAD_FONT
            c.alignment = CENTER
            c.fill = HEADER_FILL
            c.border = DAY_END if p == grid.teaching(d)[-1] else BOX
            ws.column_dimensions[get_column_letter(j)].width = 7
            col_of[(w, d, p)] = j
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 6
        ws.column_dimensions["D"].width = 28
        row = 3
        for sec in sorted(dept_sections[dept]):
            meta = ctx.section_meta[sec]
            ws.cell(row, 1, sec).font = LABEL_FONT
            ws.cell(row, 2, meta["subject"]).font = CELL_FONT
            ws.cell(row, 3, meta["grade"]).font = CELL_FONT
            ws.cell(row, 4, ", ".join(sorted(meta["teacher_names"]))).font = CELL_FONT
            for slot, ms in ctx.by_section[sec].items():
                cell = ws.cell(row, col_of[slot])
                cell.value = "\n".join(m.room or "?" for m in ms)
                cell.font = CELL_FONT
                cell.alignment = WRAP
                cell.border = BOX
            row += 1
        ws.freeze_panes = "E3"
    wb.save(path)


# -------------------------------------------------- year group timetables --
def build_year_groups(ctx: Context, path: str) -> None:
    grid = ctx.grid
    teach_union = [p for p in grid.period_order if grid.kind[p] == "teaching"]
    grades = sorted(ctx.by_grade, key=lambda g: (len(g), g))  # G6..G12
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for grade in grades:
        ws = wb.create_sheet(_sheet_name(grade, used))
        ws.cell(1, 1, f"{ctx.title} — Year Group Timetable — {grade}").font = TITLE_FONT
        c = ws.cell(2, 1, "Week / Day")
        c.font = HEAD_FONT
        c.fill = HEADER_FILL
        for j, pid in enumerate(teach_union, 2):
            c = ws.cell(2, j, pid)
            c.font = HEAD_FONT
            c.alignment = CENTER
            c.fill = HEADER_FILL
            ws.column_dimensions[get_column_letter(j)].width = 22
        ws.column_dimensions["A"].width = 14
        row = 3
        gmap = ctx.by_grade[grade]
        for week in grid.weeks:
            for day in grid.days:
                lc = ws.cell(row, 1, f"Wk {week} — {day}")
                lc.font = LABEL_FONT
                lc.border = BOX
                day_teach = set(grid.teaching(day))
                for j, pid in enumerate(teach_union, 2):
                    cell = ws.cell(row, j)
                    cell.border = BOX
                    if pid not in day_teach:
                        cell.fill = DARKGREY
                        continue
                    secs = sorted(gmap.get((week, day, pid), []))
                    if not secs:
                        continue
                    if len(secs) > 8:
                        text = f"{len(secs)} sections:\n" + "\n".join(secs[:6]) + "\n…"
                    else:
                        text = "\n".join(secs)
                    cell.value = text
                    cell.font = CELL_FONT
                    cell.alignment = WRAP
                ws.row_dimensions[row].height = 52
                row += 1
        ws.freeze_panes = "B3"
    wb.save(path)


# ------------------------------------------------------ personal grid core --
def _personal_grid(ws, ctx: Context, start_row: int, cell_fn) -> None:
    """Rows = periods (all kinds, with times); columns = (week, day).

    cell_fn(week, day, pid) -> (text, style) with style in
    {"teach", "pinned", None}. Periods absent from a day's list are dark grey.
    """
    grid = ctx.grid
    hr = start_row
    c = ws.cell(hr, 1, "Period")
    c.font = HEAD_FONT
    c.fill = HEADER_FILL
    c = ws.cell(hr, 2, "Time")
    c.font = HEAD_FONT
    c.fill = HEADER_FILL
    cols: list[tuple[str, str, int]] = []
    j = 3
    for week in grid.weeks:
        for day in grid.days:
            c = ws.cell(hr, j, f"{week} {day[:3]}")
            c.font = HEAD_FONT
            c.alignment = CENTER
            c.fill = HEADER_FILL
            c.border = DAY_END if day == grid.days[-1] else BOX
            ws.column_dimensions[get_column_letter(j)].width = 13
            cols.append((week, day, j))
            j += 1
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 11
    row = hr + 1
    for pid in grid.period_order:
        lc = ws.cell(row, 1, pid)
        lc.font = LABEL_FONT
        lc.border = BOX
        st, en = ctx.grid.times.get((grid.days[0], pid), ("", ""))
        tc = ws.cell(row, 2, f"{st}–{en}")
        tc.font = CELL_FONT
        tc.border = BOX
        is_teaching = grid.kind[pid] == "teaching"
        for week, day, j in cols:
            cell = ws.cell(row, j)
            cell.border = DAY_END if day == grid.days[-1] else BOX
            if pid not in grid.day_periods[day]:
                cell.fill = DARKGREY
                continue
            if not is_teaching:
                cell.fill = GREY
                if day == "Friday":
                    fst, fen = grid.times[(day, pid)]
                    if (fst, fen) != (st, en):
                        cell.value = f"{fst}–{fen}"
                        cell.font = PIN_FONT
                continue
            text, style = cell_fn(week, day, pid)
            if text:
                cell.value = text
                cell.alignment = WRAP
                if style == "pinned":
                    cell.font = PIN_FONT
                    cell.fill = GREY
                else:
                    cell.font = CELL_FONT
        ws.row_dimensions[row].height = 30 if is_teaching else 12
        row += 1


# ------------------------------------------------------ teacher timetables --
def build_teachers(ctx: Context, path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    n_weeks = max(len(ctx.grid.weeks), 1)
    for t in ctx.teachers:
        tid = t["id"]
        ws = wb.create_sheet(_sheet_name(t["name"], used))
        ws.cell(1, 1, ctx.title).font = TITLE_FONT
        mx = t.get("max_per_week", ctx.tier_max.get(t.get("tier"), 0))
        info = (f"{t['name']} ({tid}) — tier: {t.get('tier', '?')} — "
                f"assigned {assigned_per_week(ctx, tid):g}/wk of max {mx} — "
                f"pinned {pinned_per_week(ctx, tid):g}/wk")
        ws.cell(2, 1, info).font = HEAD_FONT
        tmeet = ctx.by_teacher.get(tid, {})
        tpin = ctx.pinned.get(tid, {})

        def cell_fn(week, day, pid, tmeet=tmeet, tpin=tpin):
            slot = (week, day, pid)
            ms = tmeet.get(slot)
            if ms:
                return ("\n+ ".join(f"{m.subject}\n{m.section}\n{m.room}"
                                    for m in ms), "teach")
            if slot in tpin:
                return (_short(tpin[slot], 40), "pinned")
            return ("", None)

        _personal_grid(ws, ctx, start_row=4, cell_fn=cell_fn)
        ws.freeze_panes = "C5"
        _ = n_weeks
    wb.save(path)


# ------------------------------------------------------ student timetables --
def build_students(ctx: Context, enrolments_path: str, out_dir: str) -> list[str]:
    en = pd.read_csv(enrolments_path, dtype=str).fillna("")
    en = en.drop_duplicates(subset=["student_id", "section_id"])
    paths = []
    for grade, gdf in en.groupby("grade", sort=True):
        wb = Workbook()
        wb.remove(wb.active)
        used: set[str] = set()
        students = (gdf.groupby(["student_id", "student_name"])["section_id"]
                    .apply(list).reset_index()
                    .sort_values("student_name"))
        for r in students.itertuples(index=False):
            sid, sname, sections = r.student_id, r.student_name, r.section_id
            slot_map: dict[tuple, list[Meeting]] = defaultdict(list)
            for sec in sections:
                for slot, ms in ctx.by_section.get(sec, {}).items():
                    slot_map[slot].extend(ms)
            tag = str(sid)[-4:]
            ws = wb.create_sheet(_sheet_name(f"{sname[:24]} {tag}", used))
            ws.cell(1, 1, ctx.title).font = TITLE_FONT
            ws.cell(2, 1, f"{sname} (ID {sid}) — {grade}").font = HEAD_FONT

            def cell_fn(week, day, pid, slot_map=slot_map):
                ms = slot_map.get((week, day, pid))
                if not ms:
                    return ("", None)
                return ("\n+ ".join(
                    f"{m.subject}\n{m.section}\n{m.teacher_names.replace('|', ', ')}"
                    f"\n{m.room}" for m in ms), "teach")

            _personal_grid(ws, ctx, start_row=4, cell_fn=cell_fn)
            ws.freeze_panes = "C5"
        p = os.path.join(out_dir, f"student_timetables_{grade}.xlsx")
        wb.save(p)
        paths.append(p)
    return paths


# -------------------------------------------------------- room utilisation --
def build_rooms(ctx: Context, path: str) -> None:
    grid = ctx.grid
    slots = grid.teaching_slots()
    cfg_rooms = ctx.cfg.get("rooms", [])
    room_type = {r["id"]: r.get("type", "") for r in cfg_rooms}
    rooms = sorted(set(room_type) | set(ctx.by_room), key=str)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid"
    ws.cell(1, 1, f"{ctx.title} — Room Usage Grid").font = TITLE_FONT
    c = ws.cell(2, 1, "Room")
    c.font = HEAD_FONT
    c.fill = HEADER_FILL
    for j, (w, d, p) in enumerate(slots, 2):
        c = ws.cell(2, j, f"{w} {d[:3]} {p}")
        c.font = HEAD_FONT
        c.alignment = CENTER
        c.fill = HEADER_FILL
        c.border = DAY_END if p == grid.teaching(d)[-1] else BOX
        ws.column_dimensions[get_column_letter(j)].width = 9
    ws.column_dimensions["A"].width = 20
    for i, room in enumerate(rooms, 3):
        lc = ws.cell(i, 1, room)
        lc.font = LABEL_FONT
        lc.border = BOX
        rmap = ctx.by_room.get(room, {})
        for j, slot in enumerate(slots, 2):
            secs = rmap.get(slot)
            if secs:
                cell = ws.cell(i, j, "\n".join(sorted(secs)))
                cell.font = CELL_FONT
                cell.alignment = WRAP
    ws.freeze_panes = "B3"

    ws2 = wb.create_sheet("Utilisation")
    ws2.cell(1, 1, f"{ctx.title} — Room Utilisation (% of teaching slots used)"
             ).font = TITLE_FONT
    heads = ["Room", "Type", "Seats", "Used slots", "Available slots", "Utilisation %"]
    for j, h in enumerate(heads, 1):
        c = ws2.cell(2, j, h)
        c.font = HEAD_FONT
        c.fill = HEADER_FILL
    seats = {r["id"]: r.get("seats", "") for r in cfg_rooms}
    avail = len(slots)
    rows = []
    for room in rooms:
        usage = len(ctx.by_room.get(room, {}))
        rows.append((room, room_type.get(room, ""), seats.get(room, ""),
                     usage, avail, round(100.0 * usage / avail, 1)))
    rows.sort(key=lambda r: -r[5])
    for i, r in enumerate(rows, 3):
        for j, v in enumerate(r, 1):
            ws2.cell(i, j, v).font = CELL_FONT
    for col, w in zip("ABCDEF", (20, 14, 7, 10, 12, 12)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A3"
    wb.save(path)


# ----------------------------------------------------------- staff loading --
def build_staff_loading(ctx: Context, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Loading"
    ws.cell(1, 1, f"{ctx.title} — Staff Loading").font = TITLE_FONT
    heads = ["Teacher", "ID", "Tier", "Max/week", "Assigned/week",
             "Pinned/week", "Headroom", "Utilisation %", "Flag"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(2, j, h)
        c.font = HEAD_FONT
        c.fill = HEADER_FILL
    for i, t in enumerate(ctx.teachers, 3):
        tid = t["id"]
        mx = t.get("max_per_week", ctx.tier_max.get(t.get("tier"), 0))
        assigned = assigned_per_week(ctx, tid)
        pin = pinned_per_week(ctx, tid)
        util = (assigned / mx) if mx else None
        if assigned > mx:
            flag, fill = "OVER", RED
        elif util is not None and util >= 0.96:
            flag, fill = "NEAR MAX", AMBER
        else:
            flag, fill = "", None
        vals = [t["name"], tid, t.get("tier", ""), mx, assigned, pin,
                round(mx - assigned, 1),
                round(util * 100, 1) if util is not None else "", flag]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(i, j, v)
            cell.font = CELL_FONT
            if fill is not None:
                cell.fill = fill
    for col, w in zip("ABCDEFGHI", (24, 8, 18, 9, 13, 12, 9, 12, 10)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    wb.save(path)


# ------------------------------------------------------------ entry point --
def generate_all(config_path: str, assignments_path: str,
                 enrolments_path: str, out_dir: str) -> list[str]:
    os.makedirs(out_dir, exist_ok=True)
    ctx = load_context(config_path, assignments_path)
    paths = []
    p = os.path.join(out_dir, "master_timetable.xlsx")
    build_master(ctx, p)
    paths.append(p)
    p = os.path.join(out_dir, "department_timetables.xlsx")
    build_departments(ctx, p)
    paths.append(p)
    p = os.path.join(out_dir, "year_group_timetables.xlsx")
    build_year_groups(ctx, p)
    paths.append(p)
    p = os.path.join(out_dir, "teacher_timetables.xlsx")
    build_teachers(ctx, p)
    paths.append(p)
    if enrolments_path:
        paths.extend(build_students(ctx, enrolments_path, out_dir))
    p = os.path.join(out_dir, "room_utilisation.xlsx")
    build_rooms(ctx, p)
    paths.append(p)
    p = os.path.join(out_dir, "staff_loading.xlsx")
    build_staff_loading(ctx, p)
    paths.append(p)
    return paths
